#!/usr/bin/env python3
"""Build dataset by adding 19 delta_* columns and gov_expected_changes to a guide xlsx.

Usage example:
  python scripts/build_dataset.py \
    --wb-zip /path/to/world_bank.zip \
    --guide-xlsx /path/to/sample_guide.xlsx \
    --mapping-csv /path/to/mapping.csv \
    --output /path/to/output.xlsx
"""

from __future__ import annotations

import argparse
import csv
import io
import zipfile
from collections import defaultdict
from statistics import mean
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


ValueMap = Dict[Tuple[str, int, str], float]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--wb-zip", required=True, help="Path to World Bank zip file")
    p.add_argument("--guide-xlsx", required=True, help="Path to guide xlsx file")
    p.add_argument("--mapping-csv", required=True, help="Path to mapping csv file")
    p.add_argument("--output", required=True, help="Output xlsx path")
    p.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    p.add_argument("--country-col", default=None, help="Country column header in guide xlsx")
    p.add_argument("--year-col", default=None, help="Year column header in guide xlsx")
    p.add_argument("--agg", choices=["mean", "sum"], default="mean", help="Aggregation for gov_expected_changes")
    return p.parse_args()


def _safe_float(x: str) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_mapping(path: str) -> List[Tuple[str, str, float]]:
    out: List[Tuple[str, str, float]] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            delta_col = (row.get("delta_col") or "").strip()
            indicator = (row.get("indicator_code") or "").strip()
            if not delta_col or not indicator:
                continue
            direction = _safe_float(row.get("direction", "1"))
            out.append((delta_col, indicator, direction if direction is not None else 1.0))
    if len(out) != 19:
        raise ValueError(f"mapping 文件需要包含 19 条映射，当前为 {len(out)} 条")
    return out


def find_data_csv_in_zip(zf: zipfile.ZipFile) -> str:
    candidates = [n for n in zf.namelist() if n.lower().endswith(".csv")]
    if not candidates:
        raise FileNotFoundError("zip 中未找到 csv 文件")

    # Prefer classic WB layout: API_*_DS2_en_csv_v2_*.csv and skip metadata files.
    good = [
        n
        for n in candidates
        if "metadata" not in n.lower()
        and ("api_" in n.lower() or "data" in n.lower())
    ]
    return good[0] if good else candidates[0]


def load_wb_values_from_zip(path: str) -> ValueMap:
    values: ValueMap = {}
    with zipfile.ZipFile(path, "r") as zf:
        data_name = find_data_csv_in_zip(zf)
        raw = zf.read(data_name)

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    # WB standard columns.
    country_col = None
    indicator_col = None
    for c in reader.fieldnames or []:
        cl = c.lower()
        if country_col is None and ("country code" == cl or "country_code" == cl.replace(" ", "_")):
            country_col = c
        if indicator_col is None and ("indicator code" == cl or "indicator_code" == cl.replace(" ", "_")):
            indicator_col = c
    if not country_col or not indicator_col:
        raise ValueError("未识别到 Country Code / Indicator Code 列")

    year_cols = [c for c in (reader.fieldnames or []) if c.strip().isdigit()]
    if not year_cols:
        raise ValueError("未识别到年份列（如 1960..2023）")

    for row in reader:
        country = (row.get(country_col) or "").strip()
        indicator = (row.get(indicator_col) or "").strip()
        if not country or not indicator:
            continue
        for yc in year_cols:
            v = _safe_float(row.get(yc, ""))
            if v is None:
                continue
            values[(country, int(yc), indicator)] = v

    return values


def _normalize_header(x: str) -> str:
    return str(x).strip().lower().replace("_", " ")


def find_header_col(header_map: Dict[str, int], preferred: Optional[str], fallbacks: Iterable[str]) -> int:
    if preferred:
        key = _normalize_header(preferred)
        if key in header_map:
            return header_map[key]
        raise ValueError(f"找不到指定列: {preferred}")
    for f in fallbacks:
        k = _normalize_header(f)
        if k in header_map:
            return header_map[k]
    raise ValueError(f"找不到列，尝试过: {', '.join(fallbacks)}")


def as_country_code(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if len(s) == 3 and s.isalpha():
        return s.upper()
    return s.upper()


def as_year(v) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def main() -> None:
    args = parse_args()
    mapping = load_mapping(args.mapping_csv)
    wb_values = load_wb_values_from_zip(args.wb_zip)

    wb = load_workbook(args.guide_xlsx)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    # header index map (normalized -> 1-based col index)
    header_map: Dict[str, int] = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        val = ws.cell(row=1, column=c).value
        if val is None:
            continue
        header_map[_normalize_header(val)] = c

    country_col_idx = find_header_col(
        header_map,
        args.country_col,
        ["country code", "country", "iso3", "iso3 code", "economy"],
    )
    year_col_idx = find_header_col(
        header_map,
        args.year_col,
        ["year", "date"],
    )

    # Ensure delta columns + gov_expected_changes exist.
    next_col = ws.max_column + 1
    delta_col_indices: Dict[str, int] = {}
    for delta_col, _indicator, _direction in mapping:
        key = _normalize_header(delta_col)
        if key in header_map:
            delta_col_indices[delta_col] = header_map[key]
        else:
            ws.cell(row=1, column=next_col, value=delta_col)
            delta_col_indices[delta_col] = next_col
            header_map[key] = next_col
            next_col += 1

    gov_col_name = "gov_expected_changes"
    gov_key = _normalize_header(gov_col_name)
    if gov_key in header_map:
        gov_col_idx = header_map[gov_key]
    else:
        ws.cell(row=1, column=next_col, value=gov_col_name)
        gov_col_idx = next_col

    for r in range(2, ws.max_row + 1):
        country = as_country_code(ws.cell(row=r, column=country_col_idx).value)
        year = as_year(ws.cell(row=r, column=year_col_idx).value)

        row_deltas: List[float] = []
        if country and year is not None:
            for delta_col, indicator, direction in mapping:
                cur = wb_values.get((country, year, indicator))
                pre = wb_values.get((country, year - 1, indicator))
                target_idx = delta_col_indices[delta_col]
                if cur is None or pre is None:
                    ws.cell(row=r, column=target_idx, value=None)
                    continue
                d = direction * (cur - pre)
                ws.cell(row=r, column=target_idx, value=d)
                row_deltas.append(d)
        else:
            for delta_col, _indicator, _direction in mapping:
                ws.cell(row=r, column=delta_col_indices[delta_col], value=None)

        if not row_deltas:
            ws.cell(row=r, column=gov_col_idx, value=None)
        elif args.agg == "sum":
            ws.cell(row=r, column=gov_col_idx, value=sum(row_deltas))
        else:
            ws.cell(row=r, column=gov_col_idx, value=mean(row_deltas))

    wb.save(args.output)
    print(f"Done. Saved: {args.output}")


if __name__ == "__main__":
    main()
